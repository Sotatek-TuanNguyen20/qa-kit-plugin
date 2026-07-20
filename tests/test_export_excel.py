import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "tools"))

from export_excel import flatten_case, group_cases, sanitize_sheet_name


def test_flatten_case_maps_basic_fields():
    case = {
        "id": "IT-LOGIN-001",
        "category": {
            "large": "Nghiệp vụ đăng nhập",
            "medium": "Xác thực password",
            "small": "Password đúng biên dưới",
        },
        "condition_vi": "Password 8 ký tự (min)",
        "precondition_vi": ["Tài khoản đã được cấp", "Trạng thái hoạt động"],
        "steps": [
            {"action_vi": "Nhập password 8 ký tự", "data": {"password": "Abcd1234"}},
            {"action_vi": "Bấm nút đăng nhập"},
        ],
        "expected_vi": "Đăng nhập thành công",
        "priority": "P1",
        "trace": ["DD-2.1"],
    }
    row = flatten_case(case)
    assert row["TC ID"] == "IT-LOGIN-001"
    assert row["Medium"] == "Xác thực password"
    assert row["Small"] == "Password đúng biên dưới"
    assert row["Title"] == "Password 8 ký tự (min)"
    assert row["Precondition"] == "Tài khoản đã được cấp\nTrạng thái hoạt động"
    assert row["Steps"] == (
        "1. Nhập password 8 ký tự (data: password=Abcd1234)\n"
        "2. Bấm nút đăng nhập"
    )
    assert row["Expected Result"] == "Đăng nhập thành công"
    assert row["Priority"] == "P1"
    assert row["Trace"] == "DD-2.1"
    assert row["Status"] == ""
    assert row["Tested By"] == ""
    assert row["Date"] == ""
    assert row["Remarks"] == ""


def test_flatten_case_handles_no_precondition_or_data():
    case = {
        "id": "IT-PAYMENT-001",
        "category": {"large": "L", "medium": "M", "small": "S"},
        "condition_vi": "c",
        "precondition_vi": [],
        "steps": [{"action_vi": "Nhập thông tin thẻ"}],
        "expected_vi": "e",
        "priority": "P2",
        "trace": [],
    }
    row = flatten_case(case)
    assert row["Precondition"] == ""
    assert row["Steps"] == "1. Nhập thông tin thẻ"
    assert row["Trace"] == ""


def test_group_cases_groups_by_module_and_category_large():
    cases_by_module = {
        "login": [
            {
                "id": "IT-LOGIN-001",
                "category": {"large": "Nghiệp vụ đăng nhập", "medium": "A", "small": "a"},
                "condition_vi": "c1", "precondition_vi": [], "steps": [],
                "expected_vi": "e1", "priority": "P1", "trace": [],
            },
            {
                "id": "IT-LOGIN-002",
                "category": {"large": "Nghiệp vụ đăng nhập", "medium": "B", "small": "b"},
                "condition_vi": "c2", "precondition_vi": [], "steps": [],
                "expected_vi": "e2", "priority": "P2", "trace": [],
            },
            {
                "id": "IT-LOGIN-003",
                "category": {"large": "Nghiệp vụ quên mật khẩu", "medium": "C", "small": "c"},
                "condition_vi": "c3", "precondition_vi": [], "steps": [],
                "expected_vi": "e3", "priority": "P3", "trace": [],
            },
        ],
    }
    groups = group_cases(cases_by_module)
    assert set(groups.keys()) == {
        ("login", "Nghiệp vụ đăng nhập"),
        ("login", "Nghiệp vụ quên mật khẩu"),
    }
    login_group = groups[("login", "Nghiệp vụ đăng nhập")]
    assert len(login_group) == 2
    assert login_group[0]["No"] == 1
    assert login_group[0]["TC ID"] == "IT-LOGIN-001"
    assert login_group[1]["No"] == 2
    assert len(groups[("login", "Nghiệp vụ quên mật khẩu")]) == 1


def test_sanitize_sheet_name_truncates_and_strips_forbidden_chars():
    name = sanitize_sheet_name("login", "Nghiệp vụ đăng nhập/đăng xuất: đầy đủ chi tiết", set())
    assert len(name) <= 31
    assert not any(c in name for c in ":\\/?*[]")


def test_sanitize_sheet_name_dedupes_on_collision_after_truncation():
    long_category = "a" * 40
    existing: set[str] = set()
    first = sanitize_sheet_name("login", long_category, existing)
    existing.add(first)
    second = sanitize_sheet_name("login", long_category, existing)
    assert first != second
    assert len(second) <= 31
    assert second not in existing


from export_excel import build_summary


def test_build_summary_counts_by_priority_and_module():
    groups = {
        ("login", "A"): [{"Priority": "P1"}, {"Priority": "P2"}],
        ("payment", "B"): [{"Priority": "P1"}, {"Priority": "P1"}, {"Priority": "P3"}],
    }
    summary = build_summary(groups)

    assert list(summary.columns) == ["Priority", "login", "payment", "Total"]

    row_p1 = summary[summary["Priority"] == "P1"].iloc[0]
    assert row_p1["login"] == 1
    assert row_p1["payment"] == 2
    assert row_p1["Total"] == 3

    row_p2 = summary[summary["Priority"] == "P2"].iloc[0]
    assert row_p2["login"] == 1
    assert row_p2["payment"] == 0

    row_p3 = summary[summary["Priority"] == "P3"].iloc[0]
    assert row_p3["login"] == 0
    assert row_p3["payment"] == 1


import openpyxl

from export_excel import write_workbook


def test_write_workbook_creates_summary_and_case_sheets(tmp_path):
    groups = {
        ("login", "Nghiệp vụ đăng nhập"): [
            {
                "No": 1, "TC ID": "IT-LOGIN-001", "Medium": "A", "Small": "a",
                "Title": "t1", "Precondition": "", "Steps": "1. Bước 1\n2. Bước 2",
                "Expected Result": "e1",
                "Priority": "P1", "Trace": "", "Status": "", "Tested By": "",
                "Date": "", "Remarks": "",
            },
        ],
    }
    summary_df = build_summary(groups)
    output_path = tmp_path / "out.xlsx"

    write_workbook(groups, summary_df, output_path)

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames[0] == "Summary"
    case_sheet_names = [n for n in wb.sheetnames if n != "Summary"]
    assert len(case_sheet_names) == 1
    sheet = wb[case_sheet_names[0]]

    assert sheet.cell(row=1, column=1).value == "No"
    assert sheet.cell(row=1, column=2).value == "TC ID"
    assert sheet.cell(row=2, column=2).value == "IT-LOGIN-001"

    header_cell = sheet.cell(row=1, column=1)
    assert header_cell.fill.start_color.rgb in ("00305496", "FF305496")
    assert header_cell.font.bold is True
    assert sheet.freeze_panes == "A2"

    steps_cell = sheet.cell(row=2, column=7)
    assert steps_cell.value == "1. Bước 1\n2. Bước 2"
    assert steps_cell.alignment.wrap_text is True
    assert steps_cell.alignment.vertical == "top"


def test_write_workbook_disambiguates_sheet_names_on_real_collision(tmp_path):
    # Two different category_large values that are distinct but share their
    # first 25 characters, so their sanitized sheet names collide only after
    # MAX_SHEET_NAME_LEN truncation — this is the case the existing_names
    # threading through write_workbook()'s loop must resolve.
    category_1 = "a" * 40
    category_2 = "a" * 39 + "b"

    # Sanity check the fixture actually reproduces a post-truncation collision
    # (and not just two names that trivially differ).
    assert sanitize_sheet_name("login", category_1, set()) == sanitize_sheet_name(
        "login", category_2, set()
    )

    groups = {
        ("login", category_1): [
            {
                "No": 1, "TC ID": "IT-LOGIN-001", "Medium": "A", "Small": "a",
                "Title": "t1", "Precondition": "", "Steps": "", "Expected Result": "e1",
                "Priority": "P1", "Trace": "", "Status": "", "Tested By": "",
                "Date": "", "Remarks": "",
            },
        ],
        ("login", category_2): [
            {
                "No": 1, "TC ID": "IT-LOGIN-002", "Medium": "B", "Small": "b",
                "Title": "t2", "Precondition": "", "Steps": "", "Expected Result": "e2",
                "Priority": "P2", "Trace": "", "Status": "", "Tested By": "",
                "Date": "", "Remarks": "",
            },
        ],
    }
    summary_df = build_summary(groups)
    output_path = tmp_path / "out.xlsx"

    write_workbook(groups, summary_df, output_path)

    wb = openpyxl.load_workbook(output_path)
    case_sheet_names = [n for n in wb.sheetnames if n != "Summary"]
    assert len(case_sheet_names) == 2
    assert case_sheet_names[0] != case_sheet_names[1]

    sheet_1 = wb[case_sheet_names[0]]
    sheet_2 = wb[case_sheet_names[1]]
    assert sheet_1.cell(row=1, column=2).value == "TC ID"
    assert sheet_1.cell(row=2, column=2).value == "IT-LOGIN-001"
    assert sheet_2.cell(row=1, column=2).value == "TC ID"
    assert sheet_2.cell(row=2, column=2).value == "IT-LOGIN-002"


import pytest

from export_excel import load_testcases, main


def test_load_testcases_raises_clear_error_when_missing(tmp_path):
    with pytest.raises(FileNotFoundError, match="testcases/missing-module.yaml"):
        load_testcases("missing-module", tmp_path)


def test_main_returns_nonzero_and_prints_error_for_missing_module(tmp_path, capsys, monkeypatch):
    monkeypatch.chdir(tmp_path)
    exit_code = main(["missing-module"])
    assert exit_code == 1
    captured = capsys.readouterr()
    assert "testcases/missing-module.yaml" in captured.err


LOGIN_YAML = """\
- id: IT-LOGIN-001
  trace: [DD-2.1]
  test_level: IT
  category:
    large: "Nghiệp vụ đăng nhập"
    medium: "Xác thực password"
    small: "Password đúng biên dưới"
  viewpoint: DATA-01
  precondition_vi: ["Tài khoản đã được cấp"]
  condition_vi: "Password 8 ký tự (min)"
  steps:
    - action_vi: "Nhập password 8 ký tự"
      data: {password: "Abcd1234"}
  expected_vi: "Đăng nhập thành công"
  evidence:
    section: "DD-2.1"
    quote: "Password từ 8 đến 32 ký tự"
    source_type: db_definition
    operator: {min: 8, min_op: ">=", max: 32, max_op: "<="}
  priority: P1
  tags: [smoke]
  automatable: true
  verify:
    - method: ui_visible
      target: "màn hình chính"
      expect: true
"""

PAYMENT_YAML = """\
- id: IT-PAYMENT-001
  trace: [DD-5.1]
  test_level: IT
  category:
    large: "Nghiệp vụ thanh toán"
    medium: "Xử lý thẻ"
    small: "Thẻ hợp lệ"
  viewpoint: DATA-02
  precondition_vi: []
  condition_vi: "Thanh toán bằng thẻ hợp lệ"
  steps:
    - action_vi: "Nhập thông tin thẻ"
  expected_vi: "Thanh toán thành công"
  evidence:
    section: "DD-5.1"
    quote: "Thẻ phải còn hạn sử dụng"
    source_type: dd
  priority: P2
  tags: []
  automatable: false
"""


def test_main_end_to_end_multi_module(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    testcases_dir = tmp_path / "testcases"
    testcases_dir.mkdir()
    (testcases_dir / "login.yaml").write_text(LOGIN_YAML, encoding="utf-8")
    (testcases_dir / "payment.yaml").write_text(PAYMENT_YAML, encoding="utf-8")

    exit_code = main(["login", "payment"])
    assert exit_code == 0

    output_path = tmp_path / "testcases-export.xlsx"
    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    assert "Summary" in wb.sheetnames
    assert any(name.startswith("login_") for name in wb.sheetnames)
    assert any(name.startswith("payment_") for name in wb.sheetnames)


def test_main_single_module_default_output_name(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    testcases_dir = tmp_path / "testcases"
    testcases_dir.mkdir()
    (testcases_dir / "login.yaml").write_text(LOGIN_YAML, encoding="utf-8")

    exit_code = main(["login"])
    assert exit_code == 0
    assert (tmp_path / "login-export.xlsx").exists()


def test_main_respects_out_flag(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    testcases_dir = tmp_path / "testcases"
    testcases_dir.mkdir()
    (testcases_dir / "login.yaml").write_text(LOGIN_YAML, encoding="utf-8")

    exit_code = main(["login", "--out", "custom.xlsx"])
    assert exit_code == 0
    assert (tmp_path / "custom.xlsx").exists()
