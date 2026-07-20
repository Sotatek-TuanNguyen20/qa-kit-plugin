"""Export testcases/<module>.yaml files into a styled Excel workbook for QC."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
import yaml
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FORBIDDEN_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
MAX_SHEET_NAME_LEN = 31

COLUMN_ORDER = [
    "No", "TC ID", "Medium", "Small", "Title", "Precondition", "Steps",
    "Expected Result", "Priority", "Trace", "Status", "Tested By", "Date", "Remarks",
]

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def flatten_case(case: dict) -> dict:
    preconditions = case.get("precondition_vi", [])
    precondition_text = "\n".join(preconditions)

    steps = case.get("steps", [])
    step_lines = []
    for i, step in enumerate(steps, start=1):
        line = f"{i}. {step['action_vi']}"
        data = step.get("data")
        if data:
            data_text = ", ".join(f"{k}={v}" for k, v in data.items())
            line += f" (data: {data_text})"
        step_lines.append(line)
    steps_text = "\n".join(step_lines)

    trace_text = ", ".join(case.get("trace", []))

    category = case.get("category", {})

    return {
        "TC ID": case["id"],
        "Medium": category.get("medium", ""),
        "Small": category.get("small", ""),
        "Title": case.get("condition_vi", ""),
        "Precondition": precondition_text,
        "Steps": steps_text,
        "Expected Result": case.get("expected_vi", ""),
        "Priority": case.get("priority", ""),
        "Trace": trace_text,
        "Status": "",
        "Tested By": "",
        "Date": "",
        "Remarks": "",
    }


def sanitize_sheet_name(module: str, category_large: str, existing_names: set[str]) -> str:
    raw = f"{module}_{category_large}"
    cleaned = FORBIDDEN_SHEET_CHARS.sub("", raw)
    base = cleaned[:MAX_SHEET_NAME_LEN]
    name = base
    suffix = 2
    while name in existing_names:
        suffix_text = f"_{suffix}"
        name = base[: MAX_SHEET_NAME_LEN - len(suffix_text)] + suffix_text
        suffix += 1
    return name


def group_cases(cases_by_module: dict) -> dict:
    groups: dict = {}
    for module, cases in cases_by_module.items():
        for case in cases:
            category_large = case.get("category", {}).get("large", "")
            key = (module, category_large)
            groups.setdefault(key, []).append(flatten_case(case))
    for rows in groups.values():
        for i, row in enumerate(rows, start=1):
            row["No"] = i
    return groups


def build_summary(groups: dict) -> pd.DataFrame:
    counts: dict = {}
    modules_order: list = []
    for (module, _category_large), rows in groups.items():
        if module not in modules_order:
            modules_order.append(module)
        for row in rows:
            priority = row.get("Priority", "")
            counts.setdefault(priority, {}).setdefault(module, 0)
            counts[priority][module] += 1

    known_order = ["P1", "P2", "P3"]
    priorities_order = [p for p in known_order if p in counts] + sorted(
        p for p in counts if p not in known_order
    )

    data = []
    for priority in priorities_order:
        row = {"Priority": priority}
        total = 0
        for module in modules_order:
            count = counts[priority].get(module, 0)
            row[module] = count
            total += count
        row["Total"] = total
        data.append(row)

    return pd.DataFrame(data, columns=["Priority"] + modules_order + ["Total"])


def _style_sheet(worksheet, num_columns: int) -> None:
    for col_idx in range(1, num_columns + 1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=num_columns):
        for cell in row:
            cell.border = THIN_BORDER

    worksheet.freeze_panes = "A2"

    for col_idx in range(1, num_columns + 1):
        column_letter = get_column_letter(col_idx)
        max_len = max(
            (
                len(str(worksheet.cell(row=r, column=col_idx).value or ""))
                for r in range(1, worksheet.max_row + 1)
            ),
            default=10,
        )
        worksheet.column_dimensions[column_letter].width = min(max_len + 2, 60)


def write_workbook(groups: dict, summary_df: pd.DataFrame, output_path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        _style_sheet(writer.sheets["Summary"], len(summary_df.columns))

        existing_names = {"Summary"}
        for (module, category_large), rows in groups.items():
            sheet_name = sanitize_sheet_name(module, category_large, existing_names)
            existing_names.add(sheet_name)
            df = pd.DataFrame(rows)[COLUMN_ORDER]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], len(COLUMN_ORDER))


def load_testcases(module: str, project_root: Path) -> list:
    path = Path(project_root) / "testcases" / f"{module}.yaml"
    if not path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy testcases/{module}.yaml — kiểm tra lại tên module."
        )
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f) or []


def main(argv: list | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Export testcases/<module>.yaml sang Excel cho QC đọc."
    )
    parser.add_argument("modules", nargs="+", help="Tên module (khớp file testcases/<module>.yaml)")
    parser.add_argument("--out", default=None, help="Đường dẫn file .xlsx output")
    args = parser.parse_args(argv)

    project_root = Path.cwd()
    cases_by_module = {}
    for module in args.modules:
        try:
            cases_by_module[module] = load_testcases(module, project_root)
        except FileNotFoundError as e:
            print(f"Lỗi: {e}", file=sys.stderr)
            return 1

    groups = group_cases(cases_by_module)
    summary_df = build_summary(groups)

    if args.out:
        output_path = Path(args.out)
    elif len(args.modules) == 1:
        output_path = Path(f"{args.modules[0]}-export.xlsx")
    else:
        output_path = Path("testcases-export.xlsx")

    write_workbook(groups, summary_df, output_path)
    print(f"Đã xuất: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
